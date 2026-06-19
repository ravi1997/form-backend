"""
services/export_service.py
Export service for CSV/Excel/PDF generation from analysis results.
"""

import os
import uuid
import pandas as pd
from datetime import datetime, timezone, timedelta
from typing import Dict, Any, List, Optional, Union
from pathlib import Path
from models.analysis import Analysis, AnalysisResult, AnalysisExport
from services.base import BaseService
from logger.unified_logger import app_logger, error_logger, audit_logger
from utils.exceptions import NotFoundError, ValidationError
import io
import csv


class ExportService(BaseService):
    """Service for exporting analysis results to various formats."""

    def __init__(self):
        super().__init__(model=AnalysisExport, schema=None)
        self.export_dir = Path(os.getenv('UPLOADS_ROOT', '/var/uploads')) / 'exports'
        self.export_dir.mkdir(parents=True, exist_ok=True)

    def create_export(
        self,
        analysis_id: str,
        organization_id: str,
        format: str,
        node_ids: List[str] = None,
        run_id: str = None,
        created_by: str = None,
        filename: str = None
    ) -> AnalysisExport:
        """Create an export job."""
        try:
            # Validate analysis exists
            analysis = Analysis.objects(
                id=analysis_id, 
                organization_id=organization_id, 
                is_deleted=False
            ).first()
            
            if not analysis:
                raise NotFoundError(f"Analysis {analysis_id} not found")
            
            # Validate format
            if format not in ['csv', 'excel', 'pdf']:
                raise ValidationError(f"Unsupported export format: {format}")
            
            # Generate filename if not provided
            if not filename:
                timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
                filename = f"{analysis.name}_{timestamp}.{format}"
            
            # Create export record
            export = AnalysisExport(
                organization_id=organization_id,
                analysis_id=analysis,
                run_id=run_id,
                format=format,
                node_ids=node_ids or [],
                file_path=str(self.export_dir / filename),
                status="queued",
                created_by=created_by,
                expires_at=datetime.now(timezone.utc) + timedelta(days=7)  # 7 days TTL
            )
            export.save()
            
            audit_logger.info(
                f"Export created: ID={export.id}, AnalysisID={analysis_id}, "
                f"Format={format}, OrgID={organization_id}"
            )
            
            return export
            
        except Exception as e:
            error_logger.error(f"Error creating export: {str(e)}", exc_info=True)
            raise

    def generate_export(self, export_id: str, organization_id: str) -> str:
        """Generate export file and return file path."""
        try:
            # Get export record
            export = self.get_export(export_id, organization_id)
            
            if export.status != "queued":
                raise ValidationError(f"Export {export_id} is not in queued status")
            
            # Update status to generating
            export.status = "generating"
            export.save()
            
            # Get analysis results
            results = self._get_export_results(export)
            
            if not results:
                raise ValidationError("No results found for export")
            
            # Generate file based on format
            if export.format == "csv":
                file_path = self._generate_csv_export(export, results)
            elif export.format == "excel":
                file_path = self._generate_excel_export(export, results)
            elif export.format == "pdf":
                file_path = self._generate_pdf_export(export, results)
            else:
                raise ValidationError(f"Unsupported export format: {export.format}")
            
            # Update export record
            export.file_path = file_path
            export.file_size_bytes = os.path.getsize(file_path)
            export.status = "ready"
            export.save()
            
            audit_logger.info(
                f"Export generated: ID={export_id}, Format={export.format}, "
                f"Size={export.file_size_bytes} bytes"
            )
            
            return file_path
            
        except Exception as e:
            error_logger.error(f"Error generating export: {str(e)}", exc_info=True)
            
            # Update export status to failed
            try:
                export.status = "failed"
                export.save()
            except Exception:
                pass
            
            raise

    def get_export(self, export_id: str, organization_id: str) -> AnalysisExport:
        """Get export by ID."""
        export = AnalysisExport.objects(
            id=export_id, 
            organization_id=organization_id, 
            is_deleted=False
        ).first()
        
        if not export:
            raise NotFoundError(f"Export {export_id} not found")
        
        return export

    def download_export(self, export_id: str, organization_id: str) -> str:
        """Get file path for download."""
        export = self.get_export(export_id, organization_id)
        
        if export.status != "ready":
            raise ValidationError(f"Export {export_id} is not ready for download")
        
        if not export.file_path or not os.path.exists(export.file_path):
            raise NotFoundError(f"Export file not found: {export.file_path}")
        
        return export.file_path

    def _get_export_results(self, export: AnalysisExport) -> List[AnalysisResult]:
        """Get analysis results for export."""
        query = AnalysisResult.objects(
            analysis_id=export.analysis_id,
            organization_id=export.organization_id
        )
        
        if export.run_id:
            query = query.filter(run_id=export.run_id)
        
        if export.node_ids:
            query = query.filter(node_id__in=export.node_ids)
        
        return list(query.order_by('node_id'))

    def _generate_csv_export(self, export: AnalysisExport, results: List[AnalysisResult]) -> str:
        """Generate CSV export file."""
        try:
            # Create export directory if it doesn't exist
            os.makedirs(os.path.dirname(export.file_path), exist_ok=True)
            
            # Process all table results
            all_rows = []
            all_columns = []
            
            for result in results:
                if result.output_type == "table" and result.data:
                    data = result.data
                    columns = [col['name'] for col in result.column_definitions]
                    
                    # Add node_id prefix to columns to avoid conflicts
                    prefixed_columns = [f"{result.node_id}_{col}" for col in columns]
                    
                    # Update all columns list
                    for col in prefixed_columns:
                        if col not in all_columns:
                            all_columns.append(col)
                    
                    # Add data with prefixed columns
                    for row in data:
                        prefixed_row = {}
                        for old_col, new_col in zip(columns, prefixed_columns):
                            prefixed_row[new_col] = row.get(old_col, '')
                        all_rows.append(prefixed_row)
            
            if not all_rows:
                raise ValidationError("No table data found for CSV export")
            
            # Write CSV file
            with open(export.file_path, 'w', newline='', encoding='utf-8') as csvfile:
                writer = csv.DictWriter(csvfile, fieldnames=all_columns)
                writer.writeheader()
                writer.writerows(all_rows)
            
            return export.file_path
            
        except Exception as e:
            error_logger.error(f"Error generating CSV export: {str(e)}", exc_info=True)
            raise

    def _generate_excel_export(self, export: AnalysisExport, results: List[AnalysisResult]) -> str:
        """Generate Excel export file."""
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            # Create export directory if it doesn't exist
            os.makedirs(os.path.dirname(export.file_path), exist_ok=True)
            
            # Create Excel workbook
            wb = Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Process each result as a separate sheet
            for result in results:
                if result.output_type == "table" and result.data:
                    # Create DataFrame
                    df = pd.DataFrame(result.data)
                    
                    # Create sheet
                    ws = wb.create_sheet(title=result.node_id[:31])  # Excel sheet name limit
                    
                    # Add column headers
                    for i, col in enumerate(result.column_definitions, 1):
                        ws.cell(row=1, column=i, value=col['label'])
                    
                    # Add data
                    for i, row in enumerate(result.data, 2):
                        for j, col in enumerate(result.column_definitions, 1):
                            value = row.get(col['name'], '')
                            ws.cell(row=i, column=j, value=value)
                    
                    # Auto-fit columns
                    for column in ws.columns:
                        max_length = 0
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = (max_length + 2) * 1.2
                        ws.column_dimensions[column[0].column_letter].width = adjusted_width
                
                elif result.output_type == "value":
                    # Create sheet for value
                    ws = wb.create_sheet(title=result.node_id[:31])
                    
                    # Add value
                    ws.cell(row=1, column=1, value="Value")
                    ws.cell(row=1, column=2, value=result.data.get('value', ''))
                    
                    # Add formatted value if available
                    if 'formatted_value' in result.data:
                        ws.cell(row=2, column=1, value="Formatted Value")
                        ws.cell(row=2, column=2, value=result.data['formatted_value'])
                
                elif result.output_type == "chart_data":
                    # Create sheet for chart data
                    ws = wb.create_sheet(title=result.node_id[:31])
                    
                    chart_data = result.data
                    labels = chart_data.get('labels', [])
                    data = chart_data.get('data', [])
                    
                    # Add headers
                    ws.cell(row=1, column=1, value="Label")
                    ws.cell(row=1, column=2, value="Value")
                    
                    # Add data
                    for i, (label, value) in enumerate(zip(labels, data), 2):
                        ws.cell(row=i, column=1, value=label)
                        ws.cell(row=i, column=2, value=value)
            
            # Save workbook
            wb.save(export.file_path)
            
            return export.file_path
            
        except ImportError:
            error_logger.error("openpyxl not installed for Excel export")
            raise ValidationError("Excel export requires openpyxl package")
            
        except Exception as e:
            error_logger.error(f"Error generating Excel export: {str(e)}", exc_info=True)
            raise

    def _generate_pdf_export(self, export: AnalysisExport, results: List[AnalysisResult]) -> str:
        """Generate PDF export file."""
        try:
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.lib import colors
            
            # Create export directory if it doesn't exist
            os.makedirs(os.path.dirname(export.file_path), exist_ok=True)
            
            # Create PDF document
            doc = SimpleDocTemplate(export.file_path, pagesize=A4)
            styles = getSampleStyleSheet()
            story = []
            
            # Add title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=30,
                alignment=1  # Center alignment
            )
            story.append(Paragraph(f"Analysis Export: {export.analysis_id.name}", title_style))
            story.append(Spacer(1, 12))
            
            # Add generation info
            info_style = ParagraphStyle(
                'Info',
                parent=styles['Normal'],
                fontSize=10,
                textColor=colors.gray
            )
            story.append(Paragraph(f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S')}", info_style))
            story.append(Paragraph(f"Format: PDF", info_style))
            story.append(Spacer(1, 20))
            
            # Process each result
            for result in results:
                # Add result title
                story.append(Paragraph(f"Node: {result.node_id}", styles['Heading2']))
                
                if result.output_type == "table" and result.data:
                    # Create table
                    table_data = []
                    
                    # Add header row
                    headers = [col['label'] for col in result.column_definitions]
                    table_data.append(headers)
                    
                    # Add data rows (limit to first 100 rows for PDF)
                    data_rows = result.data[:100]
                    for row in data_rows:
                        table_row = [str(row.get(col['name'], '')) for col in result.column_definitions]
                        table_data.append(table_row)
                    
                    # Create table
                    table = Table(table_data)
                    table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 10),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black)
                    ]))
                    
                    story.append(table)
                    story.append(Spacer(1, 20))
                    
                    # Add row count info if truncated
                    if len(result.data) > 100:
                        story.append(Paragraph(f"Showing first 100 rows of {len(result.data)} total rows.", info_style))
                        story.append(Spacer(1, 10))
                
                elif result.output_type == "value":
                    # Add value
                    value = result.data.get('value', '')
                    story.append(Paragraph(f"Value: {value}", styles['Normal']))
                    
                    if 'formatted_value' in result.data:
                        story.append(Paragraph(f"Formatted: {result.data['formatted_value']}", styles['Normal']))
                    
                    story.append(Spacer(1, 20))
                
                elif result.output_type == "chart_data":
                    # Add chart data summary
                    chart_data = result.data
                    labels = chart_data.get('labels', [])
                    data = chart_data.get('data', [])
                    
                    story.append(Paragraph(f"Chart Type: {chart_data.get('chart_type', 'Unknown')}", styles['Normal']))
                    story.append(Paragraph(f"Data Points: {len(labels)}", styles['Normal']))
                    
                    # Create summary table
                    if labels and data:
                        summary_data = [['Label', 'Value']]
                        for label, value in zip(labels[:10], data[:10]):  # Limit to first 10
                            summary_data.append([str(label), str(value)])
                        
                        summary_table = Table(summary_data)
                        summary_table.setStyle(TableStyle([
                            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                            ('FONTSIZE', (0, 0), (-1, 0), 8),
                            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                            ('GRID', (0, 0), (-1, -1), 1, colors.black)
                        ]))
                        
                        story.append(summary_table)
                        story.append(Spacer(1, 20))
                        
                        if len(labels) > 10:
                            story.append(Paragraph(f"Showing first 10 of {len(labels)} data points.", info_style))
                            story.append(Spacer(1, 10))
                
                story.append(Spacer(1, 20))
            
            # Build PDF
            doc.build(story)
            
            return export.file_path
            
        except ImportError:
            error_logger.error("reportlab not installed for PDF export")
            raise ValidationError("PDF export requires reportlab package")
            
        except Exception as e:
            error_logger.error(f"Error generating PDF export: {str(e)}", exc_info=True)
            raise

    def cleanup_expired_exports(self, days: int = 7) -> Dict[str, int]:
        """Clean up expired export files."""
        try:
            cutoff_date = datetime.now(timezone.utc) - timedelta(days=days)
            
            # Find expired exports
            expired_exports = AnalysisExport.objects(
                expires_at__lt=cutoff_date,
                is_deleted=False
            )
            
            deleted_count = 0
            deleted_files = 0
            
            for export in expired_exports:
                # Delete file if exists
                if export.file_path and os.path.exists(export.file_path):
                    try:
                        os.remove(export.file_path)
                        deleted_files += 1
                    except Exception as e:
                        error_logger.warning(f"Failed to delete export file {export.file_path}: {e}")
                
                # Mark export as deleted
                export.is_deleted = True
                export.save()
                deleted_count += 1
            
            app_logger.info(f"Cleaned up {deleted_count} expired exports, {deleted_files} files")
            
            return {
                "deleted_exports": deleted_count,
                "deleted_files": deleted_files,
                "cutoff_date": cutoff_date.isoformat()
            }
            
        except Exception as e:
            error_logger.error(f"Error cleaning up expired exports: {str(e)}", exc_info=True)
            raise


# Global service instance
export_service = ExportService()